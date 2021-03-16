# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import pandas as pd

import os
from openpyxl import load_workbook


def append_df_to_excel(filename, data, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    columns_name = []
    if 'columns_name' in to_excel_kwargs:
        columns_name = to_excel_kwargs.pop('columns_name')

    df = pd.DataFrame(data=data, columns=columns_name)
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        to_excel_kwargs.update({'header': True})
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 1,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    df = pd.DataFrame(data=data)
    if sheet_name not in writer.book.sheetnames and columns_name:
        to_excel_kwargs.update({'header': True})
        df = pd.DataFrame(data=data, columns=columns_name)


    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

class AnjukePipeline:
    def process_item(self, item, spider):
        data = [[value for x, value in item.items()]]
        sheet_name = 'test'
        for k, v in item.items():
            if k == 'small_area':
                sheet_name = v
        # df1 = pd.DataFrame(data=data)
        append_df_to_excel('./xxx.xlsx', data, sheet_name=f'{sheet_name}',
                           index=False, header=False,
                           columns_name=['位置描叙', '月租（元）', '面积（平方米）',
                                         '区域描叙', '链接地址'])

        return item

